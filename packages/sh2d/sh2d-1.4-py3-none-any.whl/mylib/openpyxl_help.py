#!/usr/bin/env python3
# -*- coding: UTF-8 -*-

import logging
from openpyxl import load_workbook, Workbook

logger = logging.getLogger("main.openpyxl")

class modifyXlsx():

    def __init__(self, excel_file_path):
        """
        :param excel_file_path: 已存在的excel路径
        """
        self.excel_file_path = excel_file_path
        try:
            self.wb = load_workbook(self.excel_file_path)
        except:
            logger.error(f'{self.excel_file_path} open fail',exc_info=True)
        self.ws = self.wb.active

    def get_names(self):
        return self.wb.get_sheet_names()

    def use_sheet(self, name):
        """
        :param name: sheet名称
        :return bool: 操作状态 True or  None
        """
        try:
            self.ws = self.wb[name]
            return True
        except:
            logger.warning(f'use_sheet {self.excel_file_path}/{name} fail',exc_info=True)
            return False
   
    def write_cell(self, name, row_no, col_no, content):
        """
        修改指定单元格内容
        :param name: sheet名称
        :param row_no: 行号 1开始
        :param col_no: 列号 1开始
        :param content: 更改后的值
        :return bool: 操作状态 True or  None
        """
        if (not isinstance(row_no, int)) or (not isinstance(col_no, int)) or (not self.use_sheet(name)):
            logger.warning(f"{self.excel_file_path}/{name}/({row_no},{col_no})row_no or col_no not int",exc_info=True)
            return
        try:
            self.ws.cell(row=row_no, column=col_no).value = content
            return True
        except:
            logger.warning(f"write_cell {self.excel_file_path}/{name}/({row_no},{col_no})/{content} fail",exc_info=True)
            return False
   
    def read_cell(self, name, row_no, col_no):
        """
        读取指定单元格内容
        :param name: sheet名称
        :param row_no: 行号 1开始
        :param col_no: 列号 1开始
        :return content: 单元格内容
        """
        if (not isinstance(row_no, int)) or (not isinstance(col_no, int)) or (not self.use_sheet(name)):
            logger.warning(f"{self.excel_file_path}/{name}/({row_no},{col_no})row_no or col_no not int",exc_info=True)
            return
        try:
            return self.ws.cell(row=row_no, column=col_no).value
        except:
            logger.warning(f"read_cell {self.excel_file_path}/{name}/({row_no},{col_no}) fail",exc_info=True)


    def write_row(self, name, row_no, row):
        """
        指定位置写入行数据
        :param name: sheet名称
        :param row_no: 第几行
        :param row: 多行内容 [1,2,3]
        """
        if (not isinstance(row_no, int)) or (not self.use_sheet(name)):
            logger.warning(f"{self.excel_file_path}/{name}/({row_no})row_no or col_no not int",exc_info=True)
            return
        try:
            for i in range(len(row)):
                self.ws.cell(row=row_no, column=i+1).value = row[i]
            return True
        except Exception as e:
            logger.warning(f"write_row {self.excel_file_path}/{name}/({row_no})/{row} fail",exc_info=True)

    def write_rows(self, name, rows, row_no=None):
        """
        追加写入多行数据
        :param name: sheet名称
        :param rows: 多行内容 [[1,2,3],[...]]
        :param row_no: 起始行数
        """
        if not self.use_sheet(name):
            return
        try:
            if row_no:
                for index, row in enumerate(rows):
                    self.write_row(name, row_no+index, row)
            else:
                for row in rows:
                    self.ws.append(row)
            return True
        except:
            logger.warning(f"write_rows {self.excel_file_path}/{name}/rows fail",exc_info=True)

    def get_rows(self, name):
        """
        获取sheet数据
        :param name: sheet名称
        :return rows: 多行内容 [[1,2,3],[...]]
        """
        if not self.use_sheet(name):
            return
        return [[cell.value for cell in row] for row in self.ws.iter_rows()]

    def save(self, excel_file_path=None):
        if excel_file_path:
            self.wb.save(excel_file_path)
        else:
            self.wb.save(self.excel_file_path)


def write_xlsx(excle_name, **tables):
    wb = Workbook()
    for name in tables.keys():
        wb.create_sheet(title=name)
    for name in wb.sheetnames:
        if name not in list(tables.keys()):
            wb.remove(wb[name])

    for name, rows in tables.items():
        if not rows:
            continue
        for row in rows:
            wb[name].append(row)
    wb.save(excle_name)



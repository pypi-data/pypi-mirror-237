import requests
import numpy
import io
import openpyxl


class Matrix(list):
  
  def remove_void(self) -> None:
    while not any(self[-1]):
      self.pop(-1)
      
  def remove_all_void(self) -> None:
    self.remove_void()
    self.reverse()
    self.remove_void()
    self.reverse()
  
  def replace_none(self,new) -> None:
    for column in self:
      for i,value in enumerate(column):
        if value is None:
          column[i] = new
    
  def remove_none(self) -> None:
    for index,column in enumerate(self):
      column = [i for i in column if i is not None]
      self[index] = column
  
  def float_to_integer(self) -> None:
    for index,column in enumerate(self):
      column = list(map(lambda x : int(x) if isinstance(x,float) else x, column))
      self[index] = column
  
  def get_indexes(self, __value) -> list:
    array = numpy.array(self)
    indexes = numpy.argwhere(array == __value).tolist()
    return indexes

  def get_index(self, __value) -> list:
    return self.get_indexes(__value)[0]

class Table():
  
  def __init__(self,__table_id : str, __worksheet : str | int):
    url = f"https://docs.google.com/spreadsheets/u/0/d/{__table_id}/export?format=xlsx&id={__table_id}"
    request = requests.get(url).content
    tmp_file = io.BytesIO(request)
    workbook = openpyxl.load_workbook(tmp_file)
    
    if isinstance(__worksheet, int):
      __worksheet = workbook.sheetnames[__worksheet]
    
    self.__worksheet = workbook[__worksheet]

  def columns(self) -> Matrix:
    columns = []
    for column in self.__worksheet.iter_cols():
      column = [cell.value for cell in column]
      columns.append(column)
    return Matrix(columns)
  
  def rows(self) -> Matrix:
    rows = []
    for row in self.__worksheet.iter_rows():
      row = [cell.value for cell in row]
      rows.append(row)
    return Matrix(rows)

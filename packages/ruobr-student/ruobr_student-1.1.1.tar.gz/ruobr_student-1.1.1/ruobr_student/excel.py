from dataclasses import dataclass
from itertools import groupby,chain
import inspect

import openpyxl
from openpyxl.styles import NamedStyle, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from datetime import datetime
from calendar import monthrange
from dateutil.relativedelta import relativedelta

from .marks import Marks
from .exceptions import DateNotFoundError
from .extra import get_student_year

def string_year(month_num : int | str , year : int | str) -> str :
  if 9 <= month_num <= 12:
    return year
  else :
    return year + 1
  


@dataclass
class _modify_journal_object:
  """
Модификцаия объекта Journal (Отсутствует параметр Subject)
  """
  mark : int
  date : datetime
  
  def __repr__(self) -> str:
    return f"{self.mark}"


@dataclass
class _subject_object:
  subject : str
  marks : list[_modify_journal_object]
  
  def __repr__(self) -> str:
    return f"{self.subject}{self.marks}"
  
  def average(self) -> float | int:
    marks = list(chain(*[i.mark for i in self.marks]))
    marks = [float(i) for i in marks if i.isdigit()]
    average = round(sum(marks) / len(marks),2)
    return f"{average:g}"
  
  def marks_row(self, days_quantity : int) -> list[str] :
    result = []
    only_dates = [int(datetime.strftime(i.date,'%d')) for i in self.marks]
    
    for day_num in range(1, days_quantity + 1):
      if day_num in only_dates:
        result.append(self.marks[only_dates.index(day_num)].mark)
      else:
        result.append('')
        
    return result


@dataclass
class _month_object:
  """
Объект месяца, содержащий информацию о каждом месяце , 
включая все оценки ,
которые были получены за этот месяц 
  """
  
  def __repr__(self) -> str:
    return f"{self.name().capitalize()} {self.data}"
  
  number : int
  quantity : int
  subjects_list : list
  data : list[_subject_object]
  
  def days_list(self) -> list[str] :
    days = [i for i in range(1,self.quantity +1)]
    while len(days) < 31:
      days.append(' ')
    return days
  
  def name(self) -> str:
    return ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль', 'август', 'сентябрь',
          'октябрь', 'ноябрь', 'декабрь'][self.number-1]
  

class _styles_object:
  
  def __init__(self, workbook : openpyxl.Workbook):
    self.wb = workbook
    
    # Автоматическое добавление стилей при иницализации объекта стилей
    # method_list = [method for method in dir(self.__class__) if method.startswith('__') is False]
    method_list = [m for m in self.__class__.__dict__.values() if inspect.isfunction(m) and m != __class__.__init__]
    for method in method_list:
      method(self)
  
  def neutral(self) -> None:
    ns = NamedStyle(name='neutral')
    ns.fill = PatternFill(fgColor='f3da0b', fill_type='solid')
    ns.font = Font(color='1f0800', name='Bahnscrift', bold=False)
    ns.alignment = Alignment(horizontal='center', vertical='center')
    self.wb.add_named_style(ns)
    
  def bad(self) -> None:
    ns = NamedStyle(name='bad')
    ns.fill = PatternFill(fgColor='ff8e7a', fill_type='solid')
    ns.font = Font(color='1f0800', name='Bahnscrift', bold=True)
    ns.alignment = Alignment(horizontal='center', vertical='center')
    self.wb.add_named_style(ns)
    
    


class ExcelTable:
  
  def __init__(self, marks : Marks, year : str | int) -> None:
    self.MARKS = marks.get_all()
    self.workbook = openpyxl.Workbook()
    self.sheet = self.workbook['Sheet']
    style = _styles_object(self.workbook)
    self.year = year
  
  def __month_generator(self):
    
    # Обрезка по году (от 20XX-20XX)
    croped_marks : list = [i for i in self.MARKS if i.date.strftime('%m.%Y') in get_student_year(self.year)]
    if not croped_marks:
      raise DateNotFoundError(f"Could not find any data for {self.year}")
    months : list = groupby(croped_marks, key = lambda x : x.date.strftime('%m'))
    

    for month, data in months:
      # Сортировка данных по уч.предметам
      data = list(data)
      data.sort(key = lambda x : x.subject)
      
      # Кол-во дней в месяце
      daysPM : int = monthrange(year = int(self.year),month=int(month))[1]
      
      # Группировка данных по уч.предметам
      lessonsDataPM = groupby(data, key=lambda y : y.subject)

      # Замена groupby объектов 
      lessonsModifyDataPM = []
      
      for subject_name , journals  in lessonsDataPM:
        journals = list(journals)
        journals = list(map(lambda x : _modify_journal_object(mark=x.mark,date=x.date),journals))
        lessonsModifyDataPM.append(_subject_object(subject = subject_name,
                                                   marks = list(journals)
                                                   ))

      # Список учебных предметов в месяце
      lessonsListPM : list = [i.subject for i in lessonsDataPM]
      
      MonthObject : _month_object =  _month_object(number = int(month),
                          data = lessonsModifyDataPM,
                          subjects_list= lessonsListPM,
                          quantity=daysPM)
      
      yield MonthObject
  
  def __recolor_bad_marks(self, bad_mark_indices : list[int], style : str) -> None : 
    for i in bad_mark_indices:
          self.sheet.cell(row = self.sheet.max_row,column=i).style = style
  
  def __fill(self) -> None:
    
    for month_object in self.__month_generator():
      header = [f"{month_object.name().capitalize()} {string_year(month_object.number,self.year)}"] + month_object.days_list() + ["Средний балл"]
      self.sheet.append(header)
      
      # Изменение шрифта для обозначения месяца
      self.sheet.cell(row = self.sheet.max_row, column=1).font = "Bahnschrift SemiBold"
      self.sheet.cell(row = self.sheet.max_row, column=1).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
      
      for subject_object in month_object.data:
        body = [subject_object.subject] +  subject_object.marks_row(31) + [subject_object.average()]
        self.sheet.append(body)
        
        # Изменение шрифта и расположения для названия предмета 
        self.sheet.cell(row = self.sheet.max_row, column=1).font = "Bahnschrift SemiLight"
        
        
        
        # Выделение плохих оценок
        two_indices = [i+1 for i, x in enumerate(body[:-1]) if '2' in x]
        self.__recolor_bad_marks(two_indices,'bad')
        
        # Выделение средних оценок
        three_indices = [i+1 for i, x in enumerate(body[:-1]) if '3' in x]
        self.__recolor_bad_marks(three_indices,'neutral')
  
  def __post_processing(self) -> None:
    # Изменение размера ячеек
    for i in range(2, self.sheet.max_column+1):
      self.sheet.column_dimensions[f'{get_column_letter(i)}'].width = 5
    
    self.sheet.column_dimensions['AG'].width = 10
    self.sheet.column_dimensions['A'].width = 75
    
    # Изменение шрифтов 
    for row in self.sheet[f'B1:AO{self.sheet.max_row}']:
      for cell in row:
        cell.font = "Bahnschrift SemiLight"
        cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

    
    # Изменение названия листа
    self.sheet.title = f'Оценки {self.year}-{self.year + 1}'
    
  def save(self, __file : str | bytes) -> None:
    # Заполнение листа данными
    self.__fill()
    # Пост-обработка
    self.__post_processing()
    
    self.workbook.save(__file)



